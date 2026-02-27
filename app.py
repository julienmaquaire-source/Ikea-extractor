from flask import Flask, render_template, request, send_file, jsonify
import pdfplumber
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

app = Flask(__name__)

def extract_text_from_pdf(pdf_bytes):
    text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text

def get_lsc_number(text):
    match = re.search(r"LSC:\s*(\d+)", text)
    return match.group(1) if match else None

def get_consignment_number(text):
    match = re.search(r"Consignment number:\s*(\d+)", text)
    return match.group(1) if match else None

def parse_orders(text):
    orders = []
    blocks = re.split(r"(?=Work order:)", text)
    for block in blocks:
        if "Sales order:" not in block:
            continue
        so_match = re.search(r"Sales order:\s*(\d+)", block)
        wo_match = re.search(r"Work order:\s*(\d+)", block)
        name_match = re.search(r"Name:\s*(.+?)\s*Delivery date:", block, re.DOTALL)
        date_match = re.search(r"Delivery date:\s*(\d{4}-\d{2}-\d{2})", block)
        postal_match = re.search(r"Postal code:\s*(\d+)", block)
        city_match = re.search(r"City:\s*(.+?)(?:\n|CDU)", block)
        if not so_match:
            continue
        name = " ".join(name_match.group(1).split()) if name_match else ""
        city = city_match.group(1).strip() if city_match else ""
        cdu_ids = re.findall(r"^(\d{9})$", block, re.MULTILINE)
        cdu_str = ", ".join(cdu_ids)
        orders.append({
            "Sales Order": so_match.group(1) if so_match else "",
            "CDU Id's": cdu_str,
            "Nom Client": name,
            "Code Postal": postal_match.group(1) if postal_match else "",
            "Ville": city,
            "Date Livraison": date_match.group(1) if date_match else "",
        })
    return orders

def generate_excel(all_orders):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Orders LSC 1670"
    columns = ["Sales Order", "CDU Id's", "Nom Client", "Code Postal", "Ville", "Date Livraison", "Consignment", "Fichier PDF"]
    header_fill = PatternFill("solid", fgColor="FFCC00")
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(all_orders, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
    col_widths = [18, 20, 25, 14, 20, 16, 20, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process():
    files = request.files.getlist("pdfs")
    target_lsc = request.form.get("lsc", "1670")
    if not files:
        return jsonify({"error": "Aucun fichier reçu"}), 400

    all_orders = []
    results = []

    for f in files:
        if not f.filename.endswith(".pdf"):
            continue
        pdf_bytes = f.read()
        text = extract_text_from_pdf(pdf_bytes)
        lsc = get_lsc_number(text)
        consignment = get_consignment_number(text)

        if lsc != target_lsc:
            results.append({"file": f.filename, "skipped": True, "lsc": lsc})
            continue

        orders = parse_orders(text)
        for o in orders:
            o["Consignment"] = consignment or ""
            o["Fichier PDF"] = f.filename
        all_orders.extend(orders)
        results.append({"file": f.filename, "skipped": False, "lsc": lsc, "count": len(orders)})

    if not all_orders:
        return jsonify({"error": f"Aucune commande LSC {target_lsc} trouvée dans les fichiers.", "results": results}), 200

    excel = generate_excel(all_orders)
    filename = f"sales_orders_lsc{target_lsc}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        excel,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

if __name__ == "__main__":
    app.run(debug=True)
